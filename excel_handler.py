from openpyxl import Workbook, load_workbook


class DataHeaders:
	NAME = "name"
	LOCATION = "location"
	WEBSITE = "website"
	PHONE = "phone"
	SOCIAL_NETWORKS = "social networks"
	EMAIL = "emails"


def count_emails_in_excel(file_path):
	wb = load_workbook(filename=file_path)
	
	for sheet in wb:
		last_column_index = sheet.max_column
		
		num_rows = sheet.max_row

		email_count = 0
		
		for cell in sheet[f"{chr(ord('A') + last_column_index - 1)}"][1:]:
			if cell.value:
				email_count += 1
		
		print(f"Sheet: {sheet}, Total Rows: {num_rows}, Email Count: {email_count} ({email_count * 100 / num_rows})")
	
	return email_count

def write_contacts_to_excel(records, filename, sheet_name):
	try:
		wb = load_workbook(filename)
		if sheet_name in wb.sheetnames:
			ws = wb[sheet_name]
		else:
			ws = wb.create_sheet(title=sheet_name)
	except FileNotFoundError:
		wb = Workbook()
		ws = wb.active
		ws.title = sheet_name

	if not ws['A1'].value:
		headers = [DataHeaders.NAME, DataHeaders.LOCATION, DataHeaders.WEBSITE, DataHeaders.PHONE,
			 		DataHeaders.SOCIAL_NETWORKS, DataHeaders.EMAIL]
		ws.append(headers)

	for record in records:
		try:
			ws.append([value for value in record.__dict__.values()])
		except ValueError as e:
			print(f"{record.__dict__} has empty dict")
	wb.save(filename)


def make_company_data_for_excel(company_data_list):
	for i in range(len(company_data_list)):
		if company_data_list[i].PHONE:
			if isinstance(company_data_list[i].PHONE, list):
				result = ""
				for phone in company_data_list[i].PHONE:
					if not phone:
						continue
					else:
						result += f"{phone},"
				company_data_list[i].PHONE = result[:-1]

		if company_data_list[i].SOCIAL_NETWORKS:
			if isinstance(company_data_list[i].SOCIAL_NETWORKS, dict):
				result = ""
				for social_network, link in company_data_list[i].SOCIAL_NETWORKS.items():
					result += f"{social_network}: {link}, "
				company_data_list[i].SOCIAL_NETWORKS = result[:-2]

		if company_data_list[i].EMAIL:
			if isinstance(company_data_list[i].EMAIL, list):
				result = ""
				for email in company_data_list[i].EMAIL:
					if not email:
						continue
					else:
						result += f"{email},"
				company_data_list[i].EMAIL = result[:-1]