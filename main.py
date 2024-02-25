import openpyxl
from bs4 import BeautifulSoup
import concurrent.futures
import requests
import re
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
import os
from dotenv import load_dotenv
from selenium.webdriver.common.keys import Keys
from facebook import login_to_facebook, get_email_from_facebook
from excel_handler import make_company_data_for_excel, write_contacts_to_excel, DataHeaders, count_emails_in_excel
from company_data import CompanyData
from data_parser import social_media, make_url, extract_data
from singleton import UsedUrlsSingleton


workbook = openpyxl.load_workbook('data.xlsx')

	
sheets = [
	"Florida-Chiropractor",
	"Florida-Carpeting+and+Flooring",
	"Florida-Moving+Company",
	"Florida-plumbers+companies",
	"Kentucky-Carpeting+and+Flooring",
	"Kentucky-Chiropractor",
	"Kentucky-Moving+Company",
	"Kentucky-plumbers+companies",
	"New York-Carpeting+and+Flooring",
	"New York-Chiropractor",
	"New York-Moving+Company",
	"New York-plumbers+companies"
]

headers = [
	"keyword",
	"location",
	"name",
	"website",
	"phone"
]



def process_company_data(row):
	new_company_records = CompanyData()
	try:
		new_company_records.set_basic_info(zip(headers, row))
	except ValueError as e:
		print(f"{e} with these headers {headers}")
		return None, None

	company_facebook_urls = []
	urls = make_url(new_company_records.WEBSITE)
	if not urls:
		return None, None
	for url in urls:
		if used_urls.if_here(url):
			continue
		collected_data = extract_data(url)

		print(f"Collected data: {collected_data}")


		if not collected_data:
			continue

		social_networks = collected_data["social networks"]
		new_company_records.set_social_networks(social_networks)

		phones = collected_data["phones"]
		if phones:
			for phone in phones:
				new_company_records.add_phone(phone)

		emails = collected_data["emails"]
		if emails:
			for email in emails:
				new_company_records.add_email(email)
		else:
			if "facebook.com" in social_networks:
				company_facebook_urls.append(social_networks["facebook.com"][0])
		used_urls.add(url)


	new_company_records.PHONE = list(set(new_company_records.PHONE))
	return new_company_records, company_facebook_urls



def main():

	for sheet in sheets:
		worksheet = workbook[sheet]
		records = []
		companies_data = {}
		
		# Process rows in parallel
		i = 0
		with concurrent.futures.ThreadPoolExecutor() as executor:
			# Start processing each row concurrently
			future_to_row = {executor.submit(process_company_data, row): row for row in worksheet.iter_rows(min_row=2, values_only=True)}
			for future in concurrent.futures.as_completed(future_to_row):
				row = future_to_row[future] 
				try:
					new_company_records, company_facebook = future.result()
					if new_company_records:
						records.append(new_company_records)
						used_urls.add(new_company_records.WEBSITE)
						companies_data[new_company_records] = company_facebook
				except Exception as exc:
					print(f"Error processing the page: {exc}")
				
				print(i)
				i += 1

		get_email_from_facebook(companies_data)
		for i in range(len(records)):
			records[i].make_for_excel()
		write_contacts_to_excel(records, "answer.xlsx", sheet)


if __name__ == "__main__":
	load_dotenv()
	
	used_urls = UsedUrlsSingleton()

	main()
	
	count_emails_in_excel("answer.xlsx")